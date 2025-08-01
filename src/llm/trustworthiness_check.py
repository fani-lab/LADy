import random
import re
import pandas as pd
import os
from glob import glob

import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage


class TrustWorthiness_PromptBuilder:
    def __init__(self, task_description=None):
        self.task_description = task_description or ( "For the Given review, is the  aspect mentioned in the review? Answer Yes or No. ")

    def build_prompt(self, review_entry: dict, aspect) -> str:
        review_text = ' '.join(review_entry['sentences'][0])
        prompt = (
            f"Review: \"{review_text}\"\n"
            f"Aspect: \"{aspect}\"\n"
            f"Task: {self.task_description}\n"
            f"Output Format: {{\"Answer\": \"Yes or No\"}}\n"
            f"Answer:" )
        
        return prompt

class model_Evaluator:
    def __init__(self): pass
    def evaluator(self, path):
        result = []
        #filepath = path
        files = [f for f in glob(path + '/*.xlsx') if not os.path.basename(f).startswith('~$')]

        for i in files:
            filepath = i
            try: df = pd.read_excel(filepath, engine='openpyxl')
            except Exception as e:
                print(f"Failed to read {filepath}: {e}")
                continue
            
            total = len(df)
            correct = df['is_correct'].sum()
            accuracy = correct/total
            
            # Create confusion matrix components
            TP = ((df['expected_answer'] == 'Yes') & (df['model_prediction'] == 'Yes')).sum()
            FN = ((df['expected_answer'] == 'Yes') & (df['model_prediction'] == 'No')).sum()
            FP = ((df['expected_answer'] == 'No') & (df['model_prediction'] == 'Yes')).sum()
            TN = ((df['expected_answer'] == 'No') & (df['model_prediction'] == 'No')).sum()
            
            # Create a subfolder for plots in the current directory
            plot_dir = os.path.join(os.path.dirname(filepath), "plots")
            os.makedirs(plot_dir, exist_ok=True)

            # Define plot file paths inside that folder
            base_name = os.path.splitext(os.path.basename(filepath))[0]
            cm_path = os.path.join(plot_dir, f"{base_name}_conf_matrix.png")
            bar_path = os.path.join(plot_dir, f"{base_name}_bar_chart.png")
            pie_path = os.path.join(plot_dir, f"{base_name}_pie_chart.png")

            
            # Plot confusion matrix
            conf_matrix = pd.DataFrame([[TP, FN], [FP, TN]], index=['Actual Yes', 'Actual No'], columns=['Predicted Yes', 'Predicted No'])
            
            plt.figure(figsize=(6, 5))
            sns.heatmap(conf_matrix, annot=True, fmt="d", cmap="Blues")
            plt.title("Confusion Matrix")
            plt.savefig(cm_path, bbox_inches='tight')
            plt.close()

            result.append({
                "Total Cases": total,
                "Correct Predictions": correct,
                "Incorrect Predictions": total - correct,
                "Accuracy": round(accuracy *100, 2)
            })
            
            result_df = pd.DataFrame(result)
            
            # Load and update Excel
            wb = load_workbook(filepath)
            if 'Evaluation_summary' in wb.sheetnames: del wb['Evaluation_summary']
            sheet = wb.create_sheet('Evaluation_summary')

            # Write column headers
            for c_idx, col_name in enumerate(result_df.columns, start=1):
                sheet.cell(row=1, column=c_idx).value = col_name

            # Write data rows
            for r_idx, row in enumerate(result_df.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx).value = value

            # Insert images
            sheet.add_image(ExcelImage(cm_path), "E2")
            
            wb.save(filepath)
            print(f'Evaluation report with visuals appended to: {filepath}')
            
           
        
        


